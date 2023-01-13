import os
import time
import string
import random
import keyboard
import threading
from mailtm import Email
from asyncio import run
from tkinter import filedialog
from tkinter import messagebox
from selenium import webdriver
from openpyxl.styles import Font
from openpyxl import load_workbook
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options

from capmonstercloudclient import CapMonsterClient, ClientOptions
from capmonstercloudclient.requests import RecaptchaV2ProxylessRequest

workbook_changes = {}
exit_event = threading.Event()
stop_event = threading.Event()

client_options = ClientOptions(api_key='***')
cap_monster_client = CapMonsterClient(options=client_options)
recaptcha2request = RecaptchaV2ProxylessRequest(websiteUrl = "https://user.gto.ru/user/register", websiteKey = "6Lc7HbQUAAAAANiqU4LIcZYoVvLb72OdoU-mX5P5")

async def solve_captcha():
  return await cap_monster_client.solve_captcha(recaptcha2request)

def generateNewAccount(row, excelPath, drivePath, location):
  stop_event.clear()
  done = False
  def changeState():
    nonlocal done
    done = True

  def input_person_data():
    number = random.randint(100000000, 999999999)
    driver.execute_script('''
      let a; (a = (string) => {
        document.querySelector("#gto_user_register_form_surname").value = string.split(" ")[0]
        document.querySelector("#gto_user_register_form_firstname").value = string.split(" ")[1]
        document.querySelector("#gto_user_register_form_patronymic").value = string.split(" ")[2]
    
        if (string.split(" ")[2][string.split(" ")[2].length - 1] === "ч")
          document.querySelector("#gto_user_register_form_gender_0").checked = true
        else
          document.querySelector("#gto_user_register_form_gender_1").checked = true

        document.querySelector("#gto_user_register_form_address_actual_full_address").value = "%s"
        document.querySelector("#gto_user_register_form_phone").value = "+79%s"
        document.querySelector("#gto_user_register_form_password").value = "%s"
        document.querySelector("#gto_user_register_form_agree_personal_data").click()
      })("%s %s %s")
    ''' % (location, str(number), pswd, firstname, surname, dogname))
    address_input = driver.find_element(By.CSS_SELECTOR, "#gto_user_register_form_address_actual_full_address")
    year_input = driver.find_element(By.CSS_SELECTOR, "#gto_user_register_form_birthday")
    driver.execute_script("arguments[0].select", year_input)
    for i in range(20):
      year_input.send_keys(Keys.DELETE)
    year_input.send_keys(f"{day}.{month}.{year}")
    driver.execute_script("arguments[0].select()", address_input)
    driver.execute_script("arguments[0].blur()", address_input)
    print('\nНачалось прохождение reCAPTCHA')
    response_json = run(solve_captcha())
    if response_json['gRecaptchaResponse'] != '':
      print('\nСекретный ключ к reCAPTCHA получен')
      response = response_json['gRecaptchaResponse']
      print(response)
      driver.execute_script('document.querySelector(".g-recaptcha-response").style.display=""')
      driver.execute_script('document.querySelector(".g-recaptcha-response").innerHTML = arguments[0]', response)
      print('\nreCAPTCHA пройдена')
      driver.execute_script('arguments[0].click()', driver.find_element(By.XPATH, '//*[@id="reg-end"]'))
    else:
      print('Не удалось пройти reCAPTCHA, пожалуйста пройдите ее самостоятельно')
    

    

  def listener(message):
    subject = message['subject']
    text = message['text']

    if subject == 'Ваш код активации для сайта gto.ru':
      index = text.find('gto.ru: ')
      confirmation_code = text[index + 7 : index + 14]

      code_input = driver.find_element(By.CSS_SELECTOR, '#gto_user_register_form_activateCode')
      confirm_button = driver.find_element(By.CSS_SELECTOR, 'button[name="next"]')

      driver.execute_script(f"arguments[0].value = '{confirmation_code}'", code_input)
      driver.execute_script("arguments[0].click()", confirm_button)
      driver.find_element(By.CSS_SELECTOR, '#gto_user_register_form_surname')
      input_person_data()
    elif subject == 'ВФСК ГТО: Ваш аккаунт активирован':
      index = text.find('Ваш УИН: ')
      uin = text[index + 8 : index + 22]
      workbook_changes[f'E{row}'] = uin
      driver.close()
      changeState()
      # email.stop()  не работает почему-то, впадлу разбираться
      return
  
  email = Email()
  email.register()
  address = str(email.address)

  options = Options()
  options.add_argument('--lang=en')
  driver = webdriver.Chrome(options = options, executable_path = drivePath[0:-4])
  driver.implicitly_wait(10)

  book = load_workbook(excelPath, True)
  sheet = book.active

  firstname = sheet[f'A{row}'].value
  surname = sheet[f'B{row}'].value
  dogname = sheet[f'C{row}'].value
  bdate = sheet[f'D{row}'].value

  book.close()

  day = bdate.day
  month = bdate.month
  year = bdate.year
  if day < 10:  day = "0" + str(day)
  if month < 10: month = "0" + str(month)
  if 2022 - year < 18:
    print("\n\n\n\nМЕНЬШЕ 18 ЛЕТ\n\n\n\n")
    driver.quit()
    changeState()
    return

  pswdL = ''.join(random.choice(string.ascii_letters) for i in range(4))
  pswdN = ''.join(random.choice(string.digits) for i in range(4))
  pswd = pswdL + pswdN


  driver.get('https://user.gto.ru/user/register') 

  email_input_field = driver.find_element(By.CSS_SELECTOR, '#gto_user_register_form_email')
  driver.execute_script(f"arguments[0].value = '{address}';", email_input_field)
  email_input_field.send_keys(' ')

  send_code_button = driver.find_element(By.CSS_SELECTOR, 'button[style="opacity: 1;"]')
  driver.execute_script("arguments[0].click()", send_code_button)

  email.start(listener, interval = 1)
  
  while not done and not stop_event.is_set(): continue
  driver.quit()
  return



def main():
  os.system('cls')
  location = input('\nВведите населенный пункт (заранее проверьте, чтобы сайт принимал введенную вами строку): ')

  is_rows_wrong = True
  while is_rows_wrong:
    is_rs_wrong = True
    while is_rs_wrong:
      row_start = int(input("\nВведите ряд, с которого начнется обработка: "))
      if isinstance(row_start, int):
        if row_start > 0:
          is_rs_wrong = False
        else: messagebox.showerror('Ошибка', 'Введен неверный тип переменной (целое чисто больше нуля)')
      else: messagebox.showerror('Ошибка', 'Введен неверный тип переменной (целое чисто больше нуля)')

    is_re_wrong = True
    while is_re_wrong:
      row_end = int(input("\nВведите ряд, на котором она закончится: "))
      if isinstance(row_end, int):
        if row_end > 0:
          is_re_wrong = False
        else: messagebox.showerror('Ошибка', 'Введен неверный тип переменной (целое чисто больше нуля)')
      else: messagebox.showerror('Ошибка', 'Введен неверный тип переменной (целое число больше нуля)')
    
    if row_start <= row_end: is_rows_wrong = False
    else:  
      messagebox.showerror('Ошибка', 'Начальный ряд должен быть меньше конечного')
      continue

  is_threads_wrong = True
  while is_threads_wrong:
    print('\nКоличество окон было кратно количеству обрабатываемых строк (помянем, если кол-во строк будет простым числом)')
    threads = int(input("\nВведите количество окон, запущенных одновременно (не рекомендую ставить выше 4, может начать подвисать): "))
    if isinstance(threads, int):
      if threads > 0 and threads < 11: pass
    else: messagebox.showerror('Ошибка', 'Введен неверный тип переменной (целое чисто от 1 до 10)')

    if (row_end - row_start + 1) % threads == 0 or threads == 1:
      is_threads_wrong = False
    else: messagebox.showerror('Ошибка', f'Количество строк {row_end - row_start + 1} должно быть кратно {threads}!')

  is_excel_wrong = True
  while is_excel_wrong:
    print("\nУкажите путь к xlsx файлу")
    excel_filename = filedialog.askopenfilename(initialdir = "./", title = "Выбрать файл", filetypes = (("Excel файлы", ("*.xlsx*", '*.xlsm', '*.xltx', '*.xltm')),("all files", "*.*")))
    if excel_filename.endswith(".xlsx") or excel_filename.endswith(".xlsm") or excel_filename.endswith(".xltx") or excel_filename.endswith(".xltm"): is_excel_wrong = False
    else: messagebox.showerror('Ошибка', 'Выбран файл не с расширением Excel (расширения .xlsx, .slsm, .xltx, .xltm)')

  is_drive_wrong = True
  while is_drive_wrong:
    print("\nУкажите путь к chromedriver файлу")
    driver_filename = filedialog.askopenfilename(initialdir = "./", title = "Выбрать файл", filetypes = (("Приложения", "*.exe*"),("all files", "*.*")))
    if driver_filename.endswith(".exe"): is_drive_wrong = False
    else: messagebox.showerror('Ошибка', 'Выбран неверный файл (не с расширением .exe)')

  threading.Thread(target = cancelInputListener).start()
  print('\nВЫ МОЖЕТЕ НАЖАТЬ Q, ЧТОБЫ ЗАКРЫТЬ ПРОГРАММУ ДОСРОЧНО\n\n\n\n')

  success = False
  for i in range(row_start, row_end + 1, threads):
    ths = [threading.Thread(target = generateNewAccount, args = (i + j, excel_filename, driver_filename, location), daemon = True) for j in range(threads)]
    for th in ths:
      th.start()
    print('Открыт поток регистрации')
    for th in ths:
      th.join()
    print('Закрыт поток регистрации')

    if i == row_end or i + threads - 1 == row_end: success = True
    if exit_event.is_set():
      break
    if i + threads - 1 < row_end:
      continue
    else: break

  if not exit_event.is_set(): exit_event.set()

  print('Внесение и сохранение изменений в таблицу Excel')
  book = load_workbook(excel_filename)
  sheet = book.active
  for row in workbook_changes:
    sheet[row].value = workbook_changes[row]
  book.save(excel_filename)
  book.close()

  print('\nВсе изменения внесены')

  if success:
    input('\n\nПрограмма завершила свою работу успешно. Нажмите любую клавишу, чтобы продолжить\n')
  else:
    input('\n\nПрограмма завершила свою работу досрочно. Нажмите любую клавишу, чтобы продолжить\n')

def cancelInputListener():
  exit_event.clear()
  while True:
    if exit_event.is_set(): return

    try:
      if keyboard.is_pressed('q'):
        # os.system('cls')
        print('\n\n\nБраузер закроется через момент. Подождите...')
        stop_event.set()
        exit_event.set()
        return
    except: break

while True:
  th = threading.Thread(target = main)
  th.start()
  th.join()


# Краснодарский край, Лабинский р-н, ст-ца Владимирская
